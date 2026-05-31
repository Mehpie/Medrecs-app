#!/usr/bin/env python3
"""Generate structured side-by-side Excel workbook from patent_review.json."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from patent_comparison_rows import ComparisonRow, load_comparison_rows

HEADER_FILL = PatternFill("solid", fgColor="E6EEF8")
STATUS_FILLS = {
    "CHANGED": PatternFill("solid", fgColor="FFF4CC"),
    "ADDED": PatternFill("solid", fgColor="D9F2D9"),
    "REMOVED": PatternFill("solid", fgColor="FFD9D9"),
}
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _apply_row_style(ws, row_idx: int, status: str, num_cols: int) -> None:
    fill = STATUS_FILLS.get(status)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = WRAP
        if fill:
            cell.fill = fill


def _write_sheet(
    ws,
    headers: list[str],
    rows: list[ComparisonRow],
    row_values,
) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    for row in rows:
        ws.append(row_values(row))
        row_idx = ws.max_row
        _apply_row_style(ws, row_idx, row.status, len(headers))

    widths = [22, 12, 70, 70] if len(headers) == 4 else [18, 22, 12, 70, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def build_xlsx_from_json(json_path: Path, output_xlsx: Path) -> None:
    claim_rows, section_rows = load_comparison_rows(json_path)

    wb = Workbook()
    claims_ws = wb.active
    claims_ws.title = "Claims"
    _write_sheet(
        claims_ws,
        ["ID", "Status", "Original", "Rewritten"],
        claim_rows,
        lambda r: [r.id_label, r.status, r.original, r.rewritten],
    )

    spec_ws = wb.create_sheet("Specification")
    _write_sheet(
        spec_ws,
        ["Section ID", "Section Title", "Paragraph ID", "Status", "Original", "Rewritten"],
        section_rows,
        lambda r: [
            r.section_id,
            r.section_title,
            r.id_label,
            r.status,
            r.original,
            r.rewritten,
        ],
    )
    spec_widths = [18, 28, 14, 12, 70, 70]
    for idx, width in enumerate(spec_widths, start=1):
        spec_ws.column_dimensions[get_column_letter(idx)].width = width

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate structured side-by-side Excel from patent_review.json"
    )
    parser.add_argument(
        "--json",
        default="data/outputs/patent_review.json",
        help="Path to patent_review.json",
    )
    parser.add_argument(
        "--output",
        default="data/outputs/patent_changes_side_by_side.xlsx",
        help="Output Excel path",
    )
    args = parser.parse_args()

    json_path = Path(args.json).resolve()
    output_path = Path(args.output).resolve()
    build_xlsx_from_json(json_path, output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
